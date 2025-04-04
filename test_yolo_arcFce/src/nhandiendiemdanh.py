import cv2
import os
import numpy as np
import onnxruntime
from insightface.app import FaceAnalysis
from sklearn.metrics.pairwise import cosine_similarity
import tkinter as tk
from tkinter import messagebox, simpledialog
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Cấu hình đường dẫn
EMBEDDINGS_PATH = r"C:\Users\V5030587\Downloads\pypj\Thang_may_nhan_dien_khuon_mat\Thang_may_nhan_dien_khuon_mat\test_yolo_arcFce\data\faces-embeddings.npz"
SAVE_DIR = r"C:\Users\V5030587\Downloads\pypj\Thang_may_nhan_dien_khuon_mat\Thang_may_nhan_dien_khuon_mat\test_yolo_arcFce\data\registered_faces"
ATTENDANCE_FILE = "diemdanh.xlsx"

# Khởi tạo insightface
app = FaceAnalysis(name="buffalo_l")
app.prepare(ctx_id=0)

# ==================== EMBEDDING VÀ QUẢN LÝ =====================

def load_database():
    if not os.path.exists(EMBEDDINGS_PATH):
        return [], np.array([])
    data = np.load(EMBEDDINGS_PATH, allow_pickle=True)
    return list(data.keys()), np.array(list(data.values()))

def save_embeddings(new_embeddings):
    names, embeddings = load_database()
    names.extend(new_embeddings.keys())
    embeddings = np.vstack([embeddings, list(new_embeddings.values())]) if embeddings.size else np.array(list(new_embeddings.values()))
    np.savez(EMBEDDINGS_PATH, **dict(zip(names, embeddings)))

def remove_embeddings():
    username = simpledialog.askstring("Xóa người dùng", "Nhập tên người dùng cần xóa:")
    if not username:
        return
    if not os.path.exists(EMBEDDINGS_PATH):
        messagebox.showerror("Lỗi", "Không tìm thấy dữ liệu embeddings!")
        return
    data = np.load(EMBEDDINGS_PATH, allow_pickle=True)
    data_dict = dict(data)
    to_delete = [key for key in data_dict if key.startswith(username)]
    if to_delete:
        for key in to_delete:
            data_dict.pop(key, None)
        np.savez(EMBEDDINGS_PATH, **data_dict)
        messagebox.showinfo("Thành công", f"Đã xóa dữ liệu người dùng: {username}")
    else:
        messagebox.showerror("Lỗi", "Không tìm thấy người dùng trong hệ thống!")

# ==================== CHỤP ẢNH VÀ GHI EMBEDDING =====================

def extract_embedding(frame):
    img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    faces = app.get(img)
    if not faces:
        return None, None
    return [face.normed_embedding for face in faces], [face.bbox for face in faces]

def capture_images(username):
    os.makedirs(SAVE_DIR, exist_ok=True)
    cap = cv2.VideoCapture(0)
    count = 0
    embeddings = {}
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        cv2.imshow("Capture Image", frame)
        key = cv2.waitKey(1) & 0xFF
        if key == ord('s'):
            file_path = os.path.join(SAVE_DIR, f"{username}_{count}.jpg")
            cv2.imwrite(file_path, frame)
            embedding, _ = extract_embedding(frame)
            if embedding:
                embeddings[f"{username}_{count}"] = embedding[0]
                print(f"Đã lưu: {file_path}")
            count += 1
        elif key == ord('q'):
            break
    cap.release()
    cv2.destroyAllWindows()
    if embeddings:
        save_embeddings(embeddings)

# ==================== NHẬN DIỆN & ĐIỂM DANH =====================

def recognize_face(frame, names, embeddings):
    face_embeddings, bboxes = extract_embedding(frame)
    if face_embeddings is None:
        return []
    results = []
    for emb, bbox in zip(face_embeddings, bboxes):
        similarities = cosine_similarity([emb], embeddings)[0]
        best_match_idx = np.argmax(similarities)
        best_score = similarities[best_match_idx]
        name = names[best_match_idx].split("_")[0] if best_score > 0.6 else "Unknown"
        results.append((name, bbox))
    return results

def mark_attendance(name):
    now = datetime.now()
    current_date = now.strftime("%Y-%m-%d")
    current_time = now.strftime("%H:%M:%S")

    if not os.path.exists(ATTENDANCE_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Tên", "Ngày", "Giờ"])
        wb.save(ATTENDANCE_FILE)

    wb = load_workbook(ATTENDANCE_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == name and row[1] == current_date:
            return False

    ws.append([name, current_date, current_time])
    wb.save(ATTENDANCE_FILE)
    return True

def start_recognition():
    names, embeddings = load_database()
    cap = cv2.VideoCapture(0)
    recognized_today = set()

    while True:
        ret, frame = cap.read()
        if not ret:
            break
        faces = recognize_face(frame, names, embeddings)
        for name, bbox in faces:
            x1, y1, x2, y2 = map(int, bbox)
            color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)
            cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
            cv2.putText(frame, name, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

            if name != "Unknown" and name not in recognized_today:
                if mark_attendance(name):
                    print(f"{name} đã điểm danh thành công.")
                else:
                    print(f"{name} đã điểm danh hôm nay.")
                    messagebox.showinfo("Điểm danh", f"{name} đã điểm danh hôm nay.")
                recognized_today.add(name)

        cv2.imshow("Face Recognition - Diem danh", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

# ==================== GIAO DIỆN TKINTER =====================

def login():
    username = simpledialog.askstring("Đăng ký", "Nhập tên của bạn:")
    if username:
        messagebox.showinfo("Đăng ký", f"Xin chào {username}! Nhấn 's' để chụp, 'q' để kết thúc.")
        capture_images(username)

def main():
    root = tk.Tk()
    root.title("Hệ Thống Nhận Diện Khuôn Mặt - Điểm Danh")
    root.geometry("400x350")
    
    tk.Label(root, text="Điểm Danh Khuôn Mặt", font=("Arial", 16)).pack(pady=20)
    tk.Button(root, text="Đăng ký khuôn mặt", command=login, height=2, width=25).pack(pady=10)
    tk.Button(root, text="Nhận diện & Điểm danh", command=start_recognition, height=2, width=25).pack(pady=10)
    tk.Button(root, text="Xóa người dùng", command=remove_embeddings, height=2, width=25).pack(pady=10)
    tk.Button(root, text="Thoát", command=root.quit, height=2, width=25).pack(pady=10)
    
    root.mainloop()

if __name__ == "__main__":
    main()
